import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\v4h4\Desktop\produceSales.xlsx")
sheet = wb["Sheet"]

PRICE_UPDATES = {"Garlic":3.07,
                 "Celery":1.19,
                 "Lemon":1.27}
for i in range(2, sheet.max_row):
    product_name = sheet.cell(row=i, column=1).value
    if product_name in PRICE_UPDATES:
        sheet.cell(row=i, column=2).value = PRICE_UPDATES[product_name]

wb.save(r"C:\Users\v4h4\Desktop\produceSales.xlsx")
